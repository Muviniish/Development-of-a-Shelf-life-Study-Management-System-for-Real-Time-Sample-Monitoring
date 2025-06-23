import pymysql
from tkinter import messagebox
from datetime import datetime
from openpyxl import Workbook, load_workbook

def connect_to_database():
    global mycursor, conn
    try:
        conn = pymysql.connect(host='localhost', user='root', password='abcd1234',database='approvedProducts')
        mycursor = conn.cursor()
    except:
        messagebox.showerror("Database Error", "Something went wrong while connecting to the database. Please check your connection settings.")
        return

file_name ="excel/"+ datetime.now().strftime("%Y-%m-%d") + ".xlsx"
def load():
    global wb, ws
    file_name = datetime.now().strftime("%Y-%m-%d") + ".xlsx"
    try:
        wb = load_workbook(filename=file_name)
        ws=wb.active
        ws.delete_rows(1, ws.max_row)
    except Exception as e:
        wb=Workbook()
        ws=wb.active
        ws.title="Data"

def fetch_users():
    mycursor.execute("SELECT * FROM products")
    result = mycursor.fetchall()
    return result

def search(option, value):
    mycursor.execute(f"SELECT * FROM products WHERE {option} = %s", value)
    result = mycursor.fetchall()
    return result

def update(id, name, desc, quan, email, testdate, approved):
    load()
    mycursor.execute("UPDATE products SET name = %s, description = %s, quantity = %s, test_date = %s WHERE ID = %s AND email = %s", (name, desc, quan, testdate, id, email))
    conn.commit()
    mycursor.execute("SELECT * FROM products where approved_date = %s", (approved))
    columns = [desc[0] for desc in mycursor.description]
    rows = mycursor.fetchall()
    ws.append(columns)
    for row in rows:
        ws.append(row)
    
    wb.save(file_name)

def id_exists(user_id):
    mycursor.execute("SELECT COUNT(*) FROM products WHERE ID = %s", (user_id,))
    result=mycursor.fetchone()
    return result[0]>0

def insert(user_id, name, description, quantity, email, test_date, approved):
    load()
    mycursor.execute("INSERT INTO products VALUES (%s, %s, %s, %s, %s, %s, %s)", (user_id, name, description, quantity, email, test_date, approved))
    conn.commit()
    mycursor.execute("SELECT * FROM products WHERE approved_date = %s", (approved))
    columns = [desc[0] for desc in mycursor.description]
    rows = mycursor.fetchall()
    ws.append(columns)
    for row in rows:
        ws.append(row)
    
    wb.save(file_name)

connect_to_database()